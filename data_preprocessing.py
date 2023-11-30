#功能：数据预处理，消除工频，去除坏道，全局平均重参考，数据保存为.mat和.fif格式
#作者：余江伟
#时间：2023/11/16; 2023/11/30:添加了基线校正和z-score标准化
#联系方式：15300593720

import mne
import numpy as np
import matplotlib.pyplot as plt
import json
import os
import yaml
from openpyxl import load_workbook
import hdf5storage as hdf
import pandas as pd


#reading yaml file, get parameters 读取yaml文件，获取参数
with open("D:\qq文件\交接代码\数据集\data\dataprocess.yaml", 'r',encoding='UTF-8') as f:
    data = yaml.load(f, Loader=yaml.FullLoader)

dataset = data.get('dataset')#data sources 数据来源
bad_channels_path = data.get('Badchannel_path')# bad channels file 记录坏道文件
high_freq = data.get('high_freq')#low pass filter frequcency低通滤波参数
low_freq = data.get('low_freq')#high pass filter frequency 高通滤波参数
onset_shift = data.get('onsetshift')#cropping onset shift time 裁剪onset时的偏移量
subject_num = data.get('sub_num')#subject number 数据subject数量
maxrun = data.get('maxrun')#max run numbers for subjects 每个subject最大试验次数
data_root = data.get('root')#data root path 数据根目录
bad_channel_start = data.get('bad_start')#certain data source's starting num of the bad channels file 某种数据坏道文件中记录的起点
saveroot = data.get('save_root')#processed data saving path 处理后数据的保存路径

#open bad channels file打开坏道文件
book = load_workbook(bad_channels_path)
sheet1 = book.get_sheet_by_name('Sheet1')

#choose 1 of 4 kind of data based on yaml parameter 根据参数选择某种数据来源
read_folder_path = " "
if dataset=="jh":
    read_folder_path = "/sub-jh10"
elif dataset=="ummc":
    read_folder_path = "/sub-ummc00"
elif dataset=="pt":
    read_folder_path = "/sub-pt"
elif dataset=="umf":
    read_folder_path = "/sub-umf00"

#loop for all subjects and runs 遍历该数据的所有受试者的所有实验
for i in range(1,subject_num+1):
    file_path = data_root+read_folder_path+"{}\ses-presurgery\ieeg".format(i)
    for j in range(1,maxrun+1):
        data_name = read_folder_path+"{}".format(i)+"_ses-presurgery_task-ictal_acq-ecog_run-0{}_ieeg".format(j)
        vhdr_data = file_path+data_name+".vhdr"
        json_data = file_path+data_name+".json"
        tsv_data = file_path+read_folder_path+"{}".format(i)+"_ses-presurgery_task-ictal_acq-ecog_run-0{}_events.tsv".format(j)
        #from .vhdr reading eeg file 从vhdr文件读取eeg文件
        if not(os.path.exists(vhdr_data)):
            continue
        raw = mne.io.read_raw_brainvision(vhdr_data, preload=True)
        #from .json reading powerline frequency and sampling frequceny 从json文件读取工频和采样率
        with open(json_data,'r',encoding='UTF-8') as f:
            json_info = json.load(f)
        PLF = json_info["PowerLineFrequency"]
        SF = round(json_info["SamplingFrequency"])
        #set default onset and offset 设置onset和offset的默认值
        onset = 10
        offset = int(raw.tmax)
        #for ummc data, read the annotations for onset and offset 对于ummc文件，直接读取标注获得onset和offset
        if dataset == "ummc":
            ann = raw.annotations.onset
            onset = ann[1]
            offset = ann[2]
        #for umf data, pt data and jh data, read .tsv file for onset and offset 对于umf,pt,jh数据，读取tsv文件获得onset和offset
        if dataset == "umf":
            df = pd.read_csv(tsv_data, delimiter='\t')
            for row in df.index:
                if df.loc[row]["trial_type"] == "eeg sz start":
                    onset = df.loc[row]["onset"]
                if df.loc[row]["trial_type"] == "eeg sz end":
                    offset = df.loc[row]["onset"]
                    break
        if dataset == "pt" or dataset == "jh":
            df = pd.read_csv(tsv_data, delimiter='\t')
            for row in df.index:
                if df.loc[row]["trial_type"] == "onset":
                    onset = df.loc[row]["onset"]
                if df.loc[row]["trial_type"] == "offset":
                    offset = df.loc[row]["onset"]
                    break
        raw.plot(scalings=40e-5,n_channels=129,duration=10)
        plt.show(block=True)#show raw data 展示原数据

        raw.filter(h_freq=high_freq, l_freq=low_freq)  # high-low pass filter 高-低通滤波
        raw.plot(scalings=40e-5, n_channels=129)
        plt.show(block=True)  # show data after high-low pass filter 展示高-低通滤波后的数据


        raw_notch = raw.copy().notch_filter(freqs=PLF)#removing powerline noise 去除工频噪音
        raw_notch.plot(scalings=40e-5,n_channels=129)
        plt.show(block=True)#show data after removing powerline noise 展示去除工频后的数据


        #in case there are less than 10 seconds shift before onset 避免有的run中onset前不足10秒
        processed_data = raw_notch.copy().crop(tmin=max(onset+onset_shift,0),tmax=offset)
        onset_annotation = 10 #marking the onset to the processed data 为处理后的文件记录新的onset位置
        if max(onset+onset_shift,0) == 0:
            onset_annotation = onset
        #drop bad channels based on the record 根据记录文件去除坏道
        if sheet1.cell(row = bad_channel_start+i,column = 2).value != None:
            drop_chan = sheet1.cell(row = bad_channel_start+i,column = 2).value.split(",")
            if dataset=="ummc":
                drop_chan.append('EVENT')
            processed_data = processed_data.drop_channels(drop_chan)
            processed_data.plot(scalings=40e-5)
            plt.show(block=True)#show data after cropping and dropping channels 展示截取后的数据

        #record the new parameters 记录处理后数据的参数
        ch_num = processed_data.info['nchan']
        sample_num = processed_data.times.shape[0]
        duration = round(processed_data.tmax)

        #rereferencing with common average reference  通过全局平均参考进行重参考
        rereferenced_data, ref_data = mne.set_eeg_reference(processed_data,copy=True)
        rereferenced_data.plot(scalings=40e-5)
        plt.show(block=True)#show the data after rereferencing 展示重参考后的数据

        #baseline correction 基线校正
        epochs = mne.make_fixed_length_epochs(rereferenced_data, duration=1, preload=True)
        baseline = (0, 0.999)
        epochs.apply_baseline(baseline)
        # epochs.average().plot()
        # plt.show(block=True)# Plot the average data after baseline correction 展示基线校正后的数据

        #Z-score standardization z-score标准化
        epoch_data = epochs.get_data()
        # Calculate mean and standard deviation across epochs and channels (mean or standard deviation across all epochs and time points for that specific channel.)
        mean = epoch_data.mean(axis=(0, 2), keepdims=True)
        std = epoch_data.std(axis=(0, 2), keepdims=True)
        # Apply Z-score normalization to the epochs data
        data_standardized = (epoch_data - mean) / std
        epochs_standardized = mne.EpochsArray(data_standardized, epochs.info, tmin=epochs.tmin)
        epochs_standardized.plot(scalings=40e-5)
        plt.show(block=True)# Plot the average data after Z-score standardization 展示标准化后的数据
        all_processed_data = data_standardized

        #save .fif and .mat file 保存文件为fif文件和mat文件
        save_folder_path = saveroot+"/"+dataset+"/sub{}".format(i)
        if not os.path.exists(save_folder_path):
            os.mkdir(save_folder_path)
        save_path_fif = save_folder_path + "/sub0{}_".format(i)+"run0{}_data.fif".format(j)
        save_path_mat = save_folder_path + "/sub0{}_".format(i)+"run0{}_data.mat".format(j)
        processed_data.save(save_path_fif,overwrite=True)
        hdf.savemat(file_name = save_path_mat,
                    mdict= {
                        "data":rereferenced_data.get_data(),#data array after processing 处理后的数据
                        "onset_annotation":onset_annotation,#onset annotation onset标注
                        'SamplingFrequency': SF,
                        'channel_num': ch_num,
                        'sample_num': sample_num,#采样个数
                        'duration': duration
                    })